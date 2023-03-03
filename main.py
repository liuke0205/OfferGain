import json

from openpyxl import load_workbook
import csv


def fun(categories, filepath):

    # 打开输入的xlsx文件
    wb = load_workbook(filepath + "input_file/original.xlsx")

    # 获取当前所有的sheet
    sheets = wb.worksheets

    # 按行获取每行的信息
    max_row_num = sheets[0].max_row

    # 判断输出文件是否有内容，如果没有加标头
    with open(filepath + "output_file/export_skill.csv", newline='') as csvfile:
        reader = csv.reader(csvfile)
        # 遍历csv文件
        for row in reader:
            # 如果文件有内容，则不为空
            if row:
                break
        else:
            with open(filepath + "output_file/export_skill.csv", 'a', newline='', encoding='utf-8') as csv_file:
                # 创建CSV写入对象
                csv_writer = csv.writer(csv_file)

                # 写入新的一行数据
                first_row = ["分类(必填)", "问题(必填)", "相似问题(选填-多个用##分隔)", "反例问题(选填-多个用##分隔)",
                             "机器人回答(必填-多个用##分隔)", "是否全部回复(选填-默认FALSE)", "是否停用(选填-默认FALSE)"]
                csv_writer.writerow(first_row)

    for i in range(2, max_row_num + 1):
        row_list = []
        for row in sheets[0][i]:
            row_list.append(row.value)

        question = row_list[0]
        url = row_list[1]
        dict = {
            "title": question,
            "description": "请点击查看答案",
            "url": url,
            "picurl": "https://mmbiz.qpic.cn/sz_mmbiz_jpg/NnT6wtVEINKejF1stcr5RBhSveCbpIicxxQ2yjcf6icx3xvgu53KibZNQ0ywZXsgmtU3ZyB8kicvQxjThyXdf3ia6NQ/0?wx_fmt=jpeg",
            "type": "pm",
            "authorizer_appid": "wxed370e17268cce6b",
            "public_name": "OfferGain"
        }

        news = {"news": {"articles": [dict]}}
        answer = json.dumps(news, ensure_ascii=False)
        # 格式化成输入信息
        new_row = [categories, question, "", "", answer, 'FALSE', 'FALSE']

        # 将格式化好的数据按行插入到输出文件的后面
        with open(filepath + "output_file/export_skill.csv", 'a', newline='', encoding='utf-8') as csv_file:
            # 创建CSV写入对象
            csv_writer = csv.writer(csv_file)

            # 写入新的一行数据
            csv_writer.writerow(new_row)
        csv_file.close()


if __name__ == '__main__':
    # 运行时需要
    fun("Java基础", "/Users/apple/code/python/OfferGain/")

